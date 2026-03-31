"""Module for importing tax data from various file formats."""

import json
from pathlib import Path
from typing import Any

from .types import Deduction, Income, Property, TaxDeclaration, TaxPayer


class DataImporter:
    """Import tax declaration data from various file formats."""

    @staticmethod
    def load_json(file_path: Path) -> dict[str, Any]:
        """Load JSON file.

        Args:
            file_path: Path to JSON file

        Returns:
            Parsed JSON data
        """
        with open(file_path, encoding="utf-8") as f:
            return json.load(f)

    @staticmethod
    def load_csv(file_path: Path) -> list[dict[str, str]]:
        """Load CSV file (requires csv module).

        Args:
            file_path: Path to CSV file

        Returns:
            List of dictionaries with CSV data
        """
        import csv

        data = []
        with open(file_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames is None:
                return data
            for row in reader:
                data.append(row)
        return data

    @staticmethod
    def load_excel(file_path: Path) -> dict[str, list[dict[str, Any]]]:
        """Load Excel file (requires openpyxl).

        Args:
            file_path: Path to Excel file

        Returns:
            Dictionary with sheet names as keys and data rows as values
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl required for Excel support. Install with: pip install openpyxl")

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        result = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            rows = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                rows.append(row_dict)

            result[sheet_name] = rows

        return result

    @staticmethod
    def import_personal_data(data: dict[str, Any]) -> TaxPayer:
        """Import personal data from dictionary.

        Args:
            data: Dictionary with personal data

        Returns:
            TaxPayer object
        """
        return TaxPayer(
            first_name=data["first_name"],
            last_name=data["last_name"],
            date_of_birth=data["date_of_birth"],
            tax_id=data["tax_id"],
            street=data["street"],
            city=data["city"],
            postal_code=data["postal_code"],
            email=data.get("email"),
            phone=data.get("phone"),
        )

    @staticmethod
    def import_incomes(data: list[dict[str, Any]], year: int) -> list[Income]:
        """Import incomes from list of dictionaries.

        Args:
            data: List of income records
            year: Tax year

        Returns:
            List of Income objects
        """
        incomes = []
        for record in data:
            # Handle both German and English column names
            source = record.get("source") or record.get("quelle") or record.get("Quelle", "")
            amount_str = record.get("amount") or record.get("betrag") or record.get("Betrag", "0")
            amount = float(str(amount_str).replace(",", "."))

            income = Income(
                source=source.lower().replace(" ", "_"),
                amount=amount,
                year=year,
                description=record.get("description") or record.get("beschreibung") or record.get("Beschreibung"),
            )
            incomes.append(income)

        return incomes

    @staticmethod
    def import_deductions(data: list[dict[str, Any]]) -> list[Deduction]:
        """Import deductions from list of dictionaries.

        Args:
            data: List of deduction records

        Returns:
            List of Deduction objects
        """
        deductions = []
        for record in data:
            category = record.get("category") or record.get("kategorie") or record.get("Kategorie", "")
            amount_str = record.get("amount") or record.get("betrag") or record.get("Betrag", "0")
            amount = float(str(amount_str).replace(",", "."))

            deduction = Deduction(
                category=category.lower().replace(" ", "_"),
                amount=amount,
                description=record.get("description") or record.get("beschreibung") or record.get("Beschreibung"),
            )
            deductions.append(deduction)

        return deductions

    @staticmethod
    def import_properties(data: list[dict[str, Any]]) -> list[Property]:
        """Import property data from list of dictionaries.

        Args:
            data: List of property records

        Returns:
            List of Property objects
        """
        from .types import ExpenseCategory

        properties = []
        for record in data:
            # Parse expense categories
            expenses: dict[str, float] = {}
            for key, value in record.items():
                key_lower = key.lower()
                if key_lower.startswith("expense_") or key_lower.startswith("kosten_"):
                    # Extract category name
                    category_name = (
                        key_lower.replace("expense_", "")
                        .replace("kosten_", "")
                        .upper()
                    )
                    try:
                        amount = float(str(value).replace(",", ".")) if value else 0
                        # Map to ExpenseCategory enum
                        if hasattr(ExpenseCategory, category_name):
                            expenses[category_name] = amount
                        else:
                            # Store as-is if not in enum
                            expenses[category_name] = amount
                    except (ValueError, TypeError):
                        pass

            purchase_price_str = record.get("purchase_price") or record.get("Kaufpreis") or "0"
            rental_income_str = record.get("rental_income") or record.get("Mieteinnahmen") or "0"

            purchase_price = float(str(purchase_price_str).replace(",", "."))
            rental_income = float(str(rental_income_str).replace(",", "."))

            prop = Property(
                name=record.get("name") or record.get("Name", ""),
                address=record.get("address") or record.get("Adresse", ""),
                postal_code=record.get("postal_code") or record.get("Postleitzahl", ""),
                city=record.get("city") or record.get("Stadt", ""),
                property_type=record.get("type") or record.get("Type", "apartment"),
                purchase_price=purchase_price,
                purchase_date=record.get("purchase_date") or record.get("Kaufdatum", ""),
                rental_income_monthly=rental_income,
                expenses=expenses,
                notes=record.get("notes") or record.get("Notizen"),
            )

            properties.append(prop)

        return properties

    @classmethod
    def from_directory(
        cls, directory_path: Path, year: int = 2025
    ) -> TaxDeclaration | None:
        """Import tax declaration from directory structure.

        Expected structure:
        - persoenliche_daten.json
        - einnahmen.csv or einnahmen.json
        - werbungskosten.csv or werbungskosten.json
        - immobilien.csv or immobilien.json

        Args:
            directory_path: Path to data directory
            year: Tax year

        Returns:
            TaxDeclaration object or None if not complete
        """
        directory_path = Path(directory_path)

        # Load personal data
        personal_file = directory_path / "persoenliche_daten.json"
        if not personal_file.exists():
            raise FileNotFoundError(f"Missing {personal_file}")

        personal_data = cls.load_json(personal_file)
        tax_payer = cls.import_personal_data(personal_data)

        declaration = TaxDeclaration(tax_payer=tax_payer, year=year)

        # Load incomes
        for income_file in directory_path.glob("einnahmen.*"):
            if income_file.suffix == ".json":
                data = cls.load_json(income_file)
                if isinstance(data, list):
                    declaration.incomes.extend(cls.import_incomes(data, year))
            elif income_file.suffix in [".csv"]:
                data = cls.load_csv(income_file)
                declaration.incomes.extend(cls.import_incomes(data, year))
            elif income_file.suffix in [".xlsx", ".xls"]:
                data = cls.load_excel(income_file)
                for sheet_data in data.values():
                    declaration.incomes.extend(cls.import_incomes(sheet_data, year))

        # Load deductions
        for deduction_file in directory_path.glob("werbungskosten.*"):
            if deduction_file.suffix == ".json":
                data = cls.load_json(deduction_file)
                if isinstance(data, list):
                    declaration.deductions.extend(cls.import_deductions(data))
            elif deduction_file.suffix == ".csv":
                data = cls.load_csv(deduction_file)
                declaration.deductions.extend(cls.import_deductions(data))
            elif deduction_file.suffix in [".xlsx", ".xls"]:
                data = cls.load_excel(deduction_file)
                for sheet_data in data.values():
                    declaration.deductions.extend(cls.import_deductions(sheet_data))

        # Load properties
        properties_dir = directory_path / "immobilien"
        if properties_dir.exists():
            for property_file in properties_dir.glob("*.*"):
                if property_file.is_file() and property_file.suffix in [".json", ".csv", ".xlsx", ".xls"]:
                    try:
                        if property_file.suffix == ".json":
                            data = cls.load_json(property_file)
                            if isinstance(data, list):
                                declaration.properties.extend(cls.import_properties(data))
                            else:
                                declaration.properties.extend(cls.import_properties([data]))
                        elif property_file.suffix == ".csv":
                            data = cls.load_csv(property_file)
                            declaration.properties.extend(cls.import_properties(data))
                        elif property_file.suffix in [".xlsx", ".xls"]:
                            data = cls.load_excel(property_file)
                            for sheet_data in data.values():
                                declaration.properties.extend(cls.import_properties(sheet_data))
                    except Exception as e:
                        print(f"Warning: Could not import {property_file}: {e}")

        return declaration
